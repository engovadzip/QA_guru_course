import os
import utils
import xlrd
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from zipfile import ZipFile


def test_create_archive():
    if not os.path.exists(utils.TMP_PATH):
        os.mkdir('tmp')

    with ZipFile(utils.ZIP_PATH, 'w') as hw_zip:
        os.chdir(utils.RESOURCES_PATH)
        hw_zip.write(utils.PDF_FILE)
        hw_zip.write(utils.TXT_FILE)
        hw_zip.write(utils.XLS_FILE)
        hw_zip.write(utils.XLSX_FILE)


def test_check_archive():
    hw_zip_list = ZipFile(utils.ZIP_PATH).namelist()
    assert utils.PDF_FILE in hw_zip_list
    assert utils.TXT_FILE in hw_zip_list
    assert utils.XLS_FILE in hw_zip_list
    assert utils.XLSX_FILE in hw_zip_list


def test_txt():
    txt_size = os.path.getsize(os.path.join(utils.RESOURCES_PATH, utils.TXT_FILE))

    with ZipFile(utils.ZIP_PATH) as hw_zip:
        assert txt_size == hw_zip.getinfo(utils.TXT_FILE).file_size


def test_pdf():
    with ZipFile(utils.ZIP_PATH) as hw_zip:
        res_pdf = PdfReader(os.path.join(utils.RESOURCES_PATH, utils.PDF_FILE))
        zip_pdf = PdfReader(hw_zip.open(utils.PDF_FILE))
        assert len(res_pdf.pages) == len(zip_pdf.pages)
        assert res_pdf.pages[6].extract_text() == zip_pdf.pages[6].extract_text()


def test_xls():
    with ZipFile(utils.ZIP_PATH) as hw_zip:
        with hw_zip.open(utils.XLS_FILE) as xls_zip:
            temp = xls_zip.read()
            zip_xls = xlrd.open_workbook(file_contents=temp)
        res_xls = xlrd.open_workbook(os.path.join(utils.RESOURCES_PATH, utils.XLS_FILE))
        assert zip_xls.nsheets == res_xls.nsheets
        assert zip_xls.sheet_names() == res_xls.sheet_names()

        res_sheet = res_xls.sheet_by_index(0)
        zip_sheet = zip_xls.sheet_by_index(0)
        assert res_sheet.ncols == zip_sheet.ncols
        assert res_sheet.nrows == zip_sheet.nrows


def test_xlsx():
    with ZipFile(utils.ZIP_PATH) as hw_zip:
        with hw_zip.open(utils.XLSX_FILE) as xlsx_zip:
            workbook = load_workbook(xlsx_zip)
            sheet = workbook.active
            column = sheet.cell(row=5, column=2).value
            assert column == 'Kathleen'